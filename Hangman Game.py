import random
import openpyxl
import os

filepath="C:/Users/MUHAMMAD HASNAIN/Desktop/project/data.xlsx"
if not os.path.exists(filepath):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    heading = ["Name","Status"]
    sheet.append(heading)
    workbook.save(filepath)




def hangman():
    
    list_of_words=['eduroam','hangman','pakistan','laptop']
    word=random.choice(list_of_words)
    turns=10
    guessmade=''
    valid_entry=set('abcdefghijklmnopqrstuvwxyz')
    while len(word)>0:
        main_word=''
        
        for letter in word:
            if letter in guessmade:
                main_word = main_word + letter
            else:
                main_word = main_word + '_ '
        if main_word == word:
            print(main_word)
            print('you won!!!')
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            a=[name,"Congratulation! You won"]
            sheet.append(a)
            workbook.save(filepath)
            break 
        
        print('guess the word',main_word)
        print()
        
        guess=input()
        if guess in valid_entry:
            guessmade = guessmade + guess
        else:
            print('Enter a valid character')
            guess=input()
            print()
        
        if guess not in word:
            turns=turns-1
            if turns == 9:
                print('9 turns left!!')
                print('-----------------------------')
                print('|')
                print('|')
                print('|')
                print('|_______')
                print('|       |')

            if turns == 8:
                print('8 turns left!!')
                print('-----------------------------')
                print('_____')
                print('|')
                print('|')
                print('|_________')
                print('|         |')
            
            if turns == 7:
                print('7 turns left!!')
                print('-----------------------------')
                print('___')
                print('|      O')
                print('|          ')
                print('|___________')
                print('|           |')
            
            if turns == 6:
                print('6 turns left!!')
                print('-----------------------------')
                print('___')
                print('|      O')
                print('|      |    ')
                print('|_____/_______')
                print('|             |')
            
            if turns == 5:
                print('5 turns left!!')
                print('-----------------------------')
                print('___')
                print('|      O')
                print('|      |    ')
                print('|_____/_\_____')
                print('|             |')
                
            if turns == 4:
                print('4 turns left!!')
                print('___')
                print('|     \O')
                print('|      |    ')
                print('|_____/_\_____')
                print('|             |')
                
            if turns == 3:
                print('3 turns left!!')
                print('___')
                print('|     \O/')
                print('|      |    ')
                print('|_____/_\_____')
                print('|             |')
                
            if turns == 2:
                print('2 turns left!!')
                print('-----------------------------')
                print('___')
                print('|  |  \O/')
                print('|      |    ')
                print('|_____/_\_____')
                print('|             |')

            if turns == 1:
                print('1 turns left!!! Hangman on his last breath')
                print('-----------------------------')
                print('___')
                print('|  |__\O/')
                print('|      |    ')
                print('|_____/_\_____')
                print('|             |')
                
            if turns == 0:
                print('Sorry! You loose')
                print('You let a good man die') 
                print('-----------------------------')
                print('___')
                print('|  |')
                print('|  |__O')
                print('|____/|\_____')
                print('|    /|\     |')
                
                
                
                workbook = openpyxl.load_workbook(filepath)
                sheet = workbook.active
                b=[name,"you lose"]
                sheet.append(b)
                workbook.save(filepath)
                break


while True:
    

    name = input('Enter a player name --> ')
    print('Welcome',name,'!')
    print('-------------------------')
    print('Try to guess the word in 10 attempts')
    hangman()

    user_input = input("Do you want to play again? (y/n)")
    if user_input.lower() == 'n':
        print("Thank you, Have a nice day!")
        break
    
        print("Enter a valid Charater")
